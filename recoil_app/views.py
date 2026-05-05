import shutil
from datetime import timedelta
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from .forms import CalculationForm, CompareRunsForm, MagneticBrakeForm, MagneticBrakeFormSet
from .forms import BrakeCatalogForm
from .models import (
    BrakeCatalog,
    BrakeForcePoint,
    CalculationRun,
    CalculationSnapshot,
    MagneticBrakeConfig,
)
from .services.analysis import enrich_with_basic_analysis
from .services.charting import save_interactive_charts
from .services.dynamics import RecoilParams, simulate_recoil
from .services.magnetic import (
    CurveBrakeParams,
    ForceCurvePoint,
    MagneticParams,
)
from .services.modeling import build_calculation_model
from .services.reporting import export_results_to_excel


COMPARE_CHART_FIELDS = [
    ("chart_x_t", "Перемещение от времени"),
    ("chart_v_a_t", "Скорость и ускорение от времени"),
    ("chart_v_x", "Скорость от перемещения"),
    ("chart_fmag_v", "Магнитные силы от скорости"),
    ("chart_forces_secondary", "Распределение сил от времени"),
    ("chart_x_t_recoil", "Перемещение — откат"),
    ("chart_v_a_t_recoil", "Скорость и ускорение — откат"),
    ("chart_forces_main_recoil", "Движущая и общая сила — откат"),
    ("chart_forces_secondary_recoil", "Распределение сил — откат"),
    ("chart_x_t_return", "Перемещение — накат"),
    ("chart_v_a_t_return", "Скорость и ускорение — накат"),
    ("chart_forces_secondary_return", "Распределение сил — накат"),
]


def _build_initial_from_run(run_id: str | None) -> tuple[dict, list[dict]]:
    initial_main: dict = {}
    brakes_initial: list[dict] = []

    if not run_id:
        return initial_main, brakes_initial

    try:
        source_run = CalculationRun.objects.get(pk=run_id)
    except CalculationRun.DoesNotExist:
        return initial_main, brakes_initial

    initial_main = {
        "name": f"{source_run.name}_1",
        "mass": source_run.mass,
        "angle_deg": source_run.angle_deg,
        "v0": source_run.v0,
        "x0": source_run.x0,
        "t_max": source_run.t_max,
        "dt": source_run.dt,
    }

    for brake in source_run.brakes.order_by("index"):
        brake_initial = {
            "model_type": brake.model_type,
            "name": brake.name,
            "gamma": brake.gamma,
            "delta": brake.delta,
            "xm": brake.xm,
            "ym": brake.ym,
            "dh1": brake.dh1,
            "dh2": brake.dh2,
            "dm": brake.dm,
            "n": brake.n,
            "mu": brake.mu,
            "bz": brake.bz,
            "lya": brake.lya,
            "wn0": brake.wn0,
        }

        if brake.model_type == MagneticBrakeConfig.MODEL_TYPE_CURVE:
            brake_initial["curve_source_brake_id"] = str(brake.pk)

        brakes_initial.append(brake_initial)

    return initial_main, brakes_initial


def _magnetic_params_from_cleaned_data(cleaned_data: dict) -> MagneticParams:
    return MagneticParams(
        gamma=cleaned_data["gamma"],
        delta=cleaned_data["delta"],
        xm=cleaned_data["xm"],
        ym=cleaned_data["ym"],
        dh1=cleaned_data["dh1"],
        dh2=cleaned_data["dh2"],
        dm=cleaned_data["dm"],
        n=cleaned_data["n"],
        mu=cleaned_data["mu"],
        bz=cleaned_data["bz"],
        lya=cleaned_data["lya"],
        wn0=cleaned_data["wn0"],
    )


def _curve_params_from_points(points: list[dict]) -> CurveBrakeParams:
    return CurveBrakeParams(
        points=tuple(
            ForceCurvePoint(
                velocity=point["velocity"],
                force=point["force"],
            )
            for point in points
        )
    )


def _read_chart_fragment(path: str | Path) -> str:
    return Path(path).read_text(encoding="utf-8")


def _collect_chart_html(run: CalculationRun) -> dict[str, str]:
    chart_html: dict[str, str] = {}

    for field_name, _ in COMPARE_CHART_FIELDS:
        field_file = getattr(run, field_name, None)
        if field_file:
            try:
                chart_html[field_name] = _read_chart_fragment(field_file.path)
            except FileNotFoundError:
                chart_html[field_name] = ""

    return chart_html


def _build_compare_chart_blocks(
    run_a: CalculationRun | None,
    run_b: CalculationRun | None,
) -> list[dict]:
    if not run_a or not run_b:
        return []

    chart_html_a = _collect_chart_html(run_a)
    chart_html_b = _collect_chart_html(run_b)

    blocks = []
    for field_name, title in COMPARE_CHART_FIELDS:
        html_a = chart_html_a.get(field_name, "")
        html_b = chart_html_b.get(field_name, "")
        if html_a or html_b:
            blocks.append(
                {
                    "field_name": field_name,
                    "title": title,
                    "html_a": html_a,
                    "html_b": html_b,
                }
            )
    return blocks


def _extract_snapshot_parts(run: CalculationRun) -> dict:
    phase_analysis = {}
    characteristic_points = {}
    engineering_metrics = {}

    try:
        snapshot = run.snapshot
        analysis_snapshot = snapshot.analysis_snapshot or {}
        phase_analysis = analysis_snapshot.get("phase_analysis", {})
        characteristic_points = analysis_snapshot.get("characteristic_points", {})
        engineering_metrics = analysis_snapshot.get("engineering_metrics", {})
    except CalculationSnapshot.DoesNotExist:
        pass

    return {
        "phase_analysis": phase_analysis,
        "characteristic_points": characteristic_points,
        "engineering_metrics": engineering_metrics,
    }


def _load_curve_points_from_source_brake(source_brake_id: str) -> list[dict]:
    source_brake = MagneticBrakeConfig.objects.filter(pk=source_brake_id).first()
    if source_brake is None:
        raise ValueError("Исходный тормоз для curve-характеристики не найден.")

    if source_brake.model_type != MagneticBrakeConfig.MODEL_TYPE_CURVE:
        raise ValueError("Указанный исходный тормоз не является curve-тормозом.")

    points_qs = source_brake.force_points.order_by("order", "id")
    points = [
        {
            "order": int(point.order),
            "velocity": float(point.velocity),
            "force": float(point.force),
        }
        for point in points_qs
    ]

    if len(points) < 2:
        raise ValueError(
            "У исходного curve-тормоза недостаточно точек характеристики F(v)."
        )

    return points


def _resolve_curve_sources(brake_formset) -> bool:
    ok = True

    for form in brake_formset.forms:
        if not hasattr(form, "cleaned_data"):
            continue
        if not form.cleaned_data:
            continue

        cleaned = form.cleaned_data
        if cleaned.get("model_type") != MagneticBrakeConfig.MODEL_TYPE_CURVE:
            continue

        parsed_points = cleaned.get("parsed_force_curve_points")
        if parsed_points:
            continue

        source_brake_id = (cleaned.get("curve_source_brake_id") or "").strip()
        if not source_brake_id:
            form.add_error(
                "force_curve_file",
                "Не удалось определить источник характеристики curve-тормоза.",
            )
            ok = False
            continue

        try:
            resolved_points = _load_curve_points_from_source_brake(source_brake_id)
        except ValueError as exc:
            form.add_error("force_curve_file", str(exc))
            ok = False
            continue

        cleaned["parsed_force_curve_points"] = resolved_points

    return ok


def _copy_curve_file_from_source(
    source_brake: MagneticBrakeConfig,
    target_brake: MagneticBrakeConfig,
) -> None:
    if not source_brake.curve_file:
        return

    source_brake.curve_file.open("rb")
    try:
        content = source_brake.curve_file.read()
    finally:
        source_brake.curve_file.close()

    original_name = Path(source_brake.curve_file.name).name or "curve.xlsx"
    target_brake.curve_file.save(original_name, ContentFile(content), save=True)


def _copy_curve_file_from_catalog(
    catalog_entry: BrakeCatalog,
    target_brake: MagneticBrakeConfig,
) -> None:
    """Копирует curve_file из каталога в новую запись расчёта (copy-on-use)."""
    if not catalog_entry.curve_file:
        return

    catalog_entry.curve_file.open("rb")
    try:
        content = catalog_entry.curve_file.read()
    finally:
        catalog_entry.curve_file.close()

    original_name = Path(catalog_entry.curve_file.name).name or "curve.xlsx"
    target_brake.curve_file.save(original_name, ContentFile(content), save=True)


def _create_brake_objects_and_runtime_models(
    run: CalculationRun,
    brake_formset,
) -> tuple[list[MagneticBrakeConfig], list[MagneticParams | CurveBrakeParams]]:
    brake_objects: list[MagneticBrakeConfig] = []
    runtime_brakes: list[MagneticParams | CurveBrakeParams] = []

    next_index = 1
    for form in brake_formset.forms:
        if not hasattr(form, "cleaned_data"):
            continue
        if not form.cleaned_data:
            continue

        cleaned = form.cleaned_data
        model_type = cleaned["model_type"]
        uploaded_curve_file = cleaned.get("force_curve_file")
        source_brake_id = (cleaned.get("curve_source_brake_id") or "").strip()
        catalog_source_id = cleaned.get("catalog_source_id")

        brake_obj = MagneticBrakeConfig.objects.create(
            run=run,
            index=next_index,
            model_type=model_type,
            name=cleaned.get("name", ""),
            curve_file=uploaded_curve_file if model_type == MagneticBrakeConfig.MODEL_TYPE_CURVE else None,
            gamma=cleaned.get("gamma"),
            delta=cleaned.get("delta"),
            xm=cleaned.get("xm"),
            ym=cleaned.get("ym"),
            dh1=cleaned.get("dh1"),
            dh2=cleaned.get("dh2"),
            dm=cleaned.get("dm"),
            n=cleaned.get("n"),
            mu=cleaned.get("mu"),
            bz=cleaned.get("bz"),
            lya=cleaned.get("lya"),
            wn0=cleaned.get("wn0"),
        )

        if model_type == MagneticBrakeConfig.MODEL_TYPE_CURVE:
            parsed_points = cleaned.get("parsed_force_curve_points", [])

            # Срез 3b: если выбран тормоз из каталога и файл сам не загружен —
            # копируем curve_file из каталога и парсим его в точки.
            if not uploaded_curve_file and not source_brake_id and catalog_source_id and not parsed_points:
                catalog_entry = BrakeCatalog.objects.filter(pk=catalog_source_id).first()
                if catalog_entry is not None and catalog_entry.curve_file:
                    _copy_curve_file_from_catalog(catalog_entry, brake_obj)
                    # Парсим файл в точки через тот же механизм, что и uploaded_file
                    try:
                        catalog_entry.curve_file.open("rb")
                        try:
                            from openpyxl import load_workbook  # локальный импорт
                            workbook = load_workbook(
                                catalog_entry.curve_file,
                                read_only=True,
                                data_only=True,
                            )
                            try:
                                helper_form = MagneticBrakeForm()
                                parsed_points = helper_form._parse_force_curve_sheet(
                                    workbook.active
                                )
                            finally:
                                workbook.close()
                        finally:
                            catalog_entry.curve_file.close()
                    except Exception:
                        # Если файл не парсится — оставляем пустые точки.
                        # Расчёт упадёт в _curve_params_from_points с понятной ошибкой.
                        parsed_points = []

            point_objects = [
                BrakeForcePoint(
                    brake=brake_obj,
                    order=point["order"],
                    velocity=point["velocity"],
                    force=point["force"],
                )
                for point in parsed_points
            ]
            if point_objects:
                BrakeForcePoint.objects.bulk_create(point_objects)

            if not uploaded_curve_file and source_brake_id:
                source_brake = MagneticBrakeConfig.objects.filter(pk=source_brake_id).first()
                if source_brake is not None:
                    _copy_curve_file_from_source(source_brake, brake_obj)

            runtime_brakes.append(_curve_params_from_points(parsed_points))
        else:
            runtime_brakes.append(_magnetic_params_from_cleaned_data(cleaned))

        brake_objects.append(brake_obj)
        next_index += 1

    return brake_objects, runtime_brakes


def _extract_overlay_data(run: CalculationRun) -> dict:
    """Извлекает данные timeline + ключевые индексы из result_snapshot для overlay-графиков.

    Возвращает словарь с ключами:
        t, x, v, a — массивы временных рядов
        f_magnetic — суммарная магнитная сила (по времени)
        t_recoil_end — момент разворота (для vline на графике)
        recoil_end_index — индекс точки разворота в массиве t

    Если snapshot отсутствует — возвращает пустые массивы.
    """
    out: dict = {
        "t": [], "x": [], "v": [], "a": [],
        "f_magnetic": [],
        "t_recoil_end": None,
        "recoil_end_index": None,
    }
    try:
        snap = run.snapshot
    except CalculationSnapshot.DoesNotExist:
        return out

    rs = snap.result_snapshot or {}
    timeline = rs.get("timeline") or {}
    out["t"] = timeline.get("t") or []
    out["x"] = timeline.get("x") or []
    out["v"] = timeline.get("v") or []
    out["a"] = timeline.get("a") or []

    forces = rs.get("forces") or {}
    out["f_magnetic"] = forces.get("magnetic_sum") or []

    phases = rs.get("phases") or {}
    recoil = phases.get("recoil") or {}
    out["t_recoil_end"] = recoil.get("end_time")
    out["recoil_end_index"] = recoil.get("end_index")

    return out


def _build_compare_overlay_charts(run_a: CalculationRun, run_b: CalculationRun) -> dict:
    """Генерирует 4 overlay-графика для сравнения двух расчётов."""
    from .services.charting import (
        make_compare_x_t_fragment,
        make_compare_v_a_t_fragment,
        make_compare_v_x_fragment,
        make_compare_fmag_v_fragment,
    )

    snap_a = _extract_overlay_data(run_a)
    snap_b = _extract_overlay_data(run_b)
    name_a = run_a.name or f"#{run_a.id}"
    name_b = run_b.name or f"#{run_b.id}"

    return {
        "x_t":    make_compare_x_t_fragment(snap_a, snap_b, name_a, name_b),
        "v_a_t":  make_compare_v_a_t_fragment(snap_a, snap_b, name_a, name_b),
        "v_x":    make_compare_v_x_fragment(snap_a, snap_b, name_a, name_b),
        "fmag_v": make_compare_fmag_v_fragment(snap_a, snap_b, name_a, name_b),
    }


def _build_compare_metrics_table(run_a: CalculationRun, run_b: CalculationRun) -> list[dict]:
    """Дельта-таблица KPI: список словарей {label, unit, value_a, value_b, delta_abs, delta_pct}.

    Берёт значения из CalculationRun-полей и из analysis_snapshot.
    """
    snap_a_parts = _extract_snapshot_parts(run_a)
    snap_b_parts = _extract_snapshot_parts(run_b)

    pa = snap_a_parts.get("phase_analysis", {}) or {}
    pb = snap_b_parts.get("phase_analysis", {}) or {}
    chars_a = snap_a_parts.get("characteristic_points", {}) or {}
    chars_b = snap_b_parts.get("characteristic_points", {}) or {}
    eng_a = snap_a_parts.get("engineering_metrics", {}) or {}
    eng_b = snap_b_parts.get("engineering_metrics", {}) or {}

    rec_a = (pa.get("recoil") or {}) if isinstance(pa, dict) else {}
    rec_b = (pb.get("recoil") or {}) if isinstance(pb, dict) else {}
    ret_a = (pa.get("return") or {}) if isinstance(pa, dict) else {}
    ret_b = (pb.get("return") or {}) if isinstance(pb, dict) else {}

    def _nested_value(d: dict, key: str):
        """В chars значения хранятся как {value, time}; берём value, если есть."""
        v = d.get(key)
        if isinstance(v, dict):
            return v.get("value")
        return v

    rows: list[tuple[str, str, object, object]] = [
        # (label, unit, value_a, value_b)
        ("Макс. перемещение",       "м",   run_a.x_max,                 run_b.x_max),
        ("Макс. скорость",          "м/с", _nested_value(chars_a, "v_max"), _nested_value(chars_b, "v_max")),
        ("Время отката",            "с",   run_a.recoil_end_time,       run_b.recoil_end_time),
        ("Время цикла",             "с",   run_a.return_end_time,       run_b.return_end_time),
        ("Энергия подведенная",     "Дж",  eng_a.get("energy_input_total"),  eng_b.get("energy_input_total")),
        ("Энергия рассеянная",      "Дж",  eng_a.get("energy_brake_total"),  eng_b.get("energy_brake_total")),
        ("Невязка энергобаланса",   "%",   run_a.energy_residual_pct,   run_b.energy_residual_pct),
        ("Откат: x_max",            "м",   rec_a.get("x_max"),  rec_b.get("x_max")),
        ("Откат: v_max",            "м/с", rec_a.get("v_max"),  rec_b.get("v_max")),
        ("Откат: a_max",            "м/с²", rec_a.get("a_max"), rec_b.get("a_max")),
        ("Накат: v_max",            "м/с", ret_a.get("v_max"),  ret_b.get("v_max")),
        ("Накат: a_max",            "м/с²", ret_a.get("a_max"), ret_b.get("a_max")),
    ]

    table: list[dict] = []
    for label, unit, va, vb in rows:
        try:
            fa = float(va) if va is not None else None
            fb = float(vb) if vb is not None else None
        except (TypeError, ValueError):
            fa = fb = None

        if fa is None or fb is None:
            delta_abs = None
            delta_pct = None
            direction = "none"
        else:
            delta_abs = fb - fa
            if fa != 0:
                delta_pct = (fb - fa) / abs(fa) * 100.0
            else:
                delta_pct = None
            if abs(delta_abs) < 1e-9:
                direction = "eq"
            elif delta_abs > 0:
                direction = "up"
            else:
                direction = "down"

        table.append({
            "label":     label,
            "unit":      unit,
            "value_a":   fa,
            "value_b":   fb,
            "delta_abs": delta_abs,
            "delta_pct": delta_pct,
            "direction": direction,
        })

    return table


def compare_view(request):
    form = CompareRunsForm(request.GET or None)

    run_a = None
    run_b = None
    overlay_charts = {}
    metrics_table: list[dict] = []

    if form.is_valid():
        run_a = form.cleaned_data["run_a"]
        run_b = form.cleaned_data["run_b"]
        overlay_charts = _build_compare_overlay_charts(run_a, run_b)
        metrics_table = _build_compare_metrics_table(run_a, run_b)

    return render(
        request,
        "recoil_app/compare.html",
        {
            "form": form,
            "run_a": run_a,
            "run_b": run_b,
            "overlay_charts": overlay_charts,
            "metrics_table": metrics_table,
        },
    )


def index_view(request):
    if request.method == "POST":
        form = CalculationForm(request.POST, request.FILES)
        brake_formset = MagneticBrakeFormSet(request.POST, request.FILES, prefix="brakes")

        forms_valid = form.is_valid() and brake_formset.is_valid()
        curve_sources_valid = _resolve_curve_sources(brake_formset) if forms_valid else False

        if forms_valid and curve_sources_valid:
            try:
                with transaction.atomic():
                    run = CalculationRun.objects.create(
                        name=form.cleaned_data["name"],
                        input_file=form.cleaned_data["input_file"],
                        mass=form.cleaned_data["mass"],
                        angle_deg=form.cleaned_data["angle_deg"],
                        v0=form.cleaned_data["v0"],
                        x0=form.cleaned_data["x0"],
                        t_max=form.cleaned_data["t_max"],
                        dt=form.cleaned_data["dt"],
                    )

                    brake_objects, runtime_brakes = _create_brake_objects_and_runtime_models(
                        run,
                        brake_formset,
                    )

                    recoil = RecoilParams(
                        mass=run.mass,
                        angle_deg=run.angle_deg,
                        v0=run.v0,
                        x0=run.x0,
                        t_max=run.t_max,
                        dt=run.dt,
                    )

                    result = simulate_recoil(run.input_file.path, recoil, runtime_brakes)

                    run.x_max = float(result.x.max())
                    run.v_max = float(result.v.max())
                    run.x_final = float(result.x[-1])
                    run.v_final = float(result.v[-1])
                    run.a_final = float(result.a[-1])
                    run.recoil_end_time = result.recoil_end_time
                    run.return_end_time = result.return_end_time
                    run.termination_reason = result.termination_reason
                    run.spring_out_of_range = result.spring_out_of_range
                    run.warnings_text = "\n".join(result.warnings)

                    safe_name = slugify(run.name) or f"run-{run.id}"
                    run_folder_name = f"{safe_name}_{run.id}"
                    prefix = run_folder_name

                    run_reports_dir = Path(settings.MEDIA_ROOT) / "reports" / run_folder_name
                    run_reports_dir.mkdir(parents=True, exist_ok=True)

                    chart_paths = save_interactive_charts(result, run_reports_dir, prefix=prefix)

                    field_map = {
                        "chart_x_t": "chart_x_t",
                        "chart_v_a_t": "chart_v_a_t",
                        "chart_v_x": "chart_v_x",
                        "chart_fmag_v": "chart_fmag_v",
                        "chart_forces_secondary": "chart_forces_secondary",
                        "chart_x_t_recoil": "chart_x_t_recoil",
                        "chart_v_a_t_recoil": "chart_v_a_t_recoil",
                        "chart_forces_main_recoil": "chart_forces_main_recoil",
                        "chart_forces_secondary_recoil": "chart_forces_secondary_recoil",
                        "chart_x_t_return": "chart_x_t_return",
                        "chart_v_a_t_return": "chart_v_a_t_return",
                        "chart_forces_secondary_return": "chart_forces_secondary_return",
                        # --- v2 ---
                        "chart_x_t_annotated": "chart_x_t_annotated",
                        "chart_energy": "chart_energy",
                    }

                    for chart_key, model_field in field_map.items():
                        if chart_key in chart_paths:
                            setattr(
                                run,
                                model_field,
                                f"reports/{run_folder_name}/{Path(chart_paths[chart_key]).name}",
                            )

                    # --- v2: энергобаланс ---
                    if result.energy_residual_pct is not None:
                        run.energy_residual_pct = float(result.energy_residual_pct)
                    if result.energy_input_cum is not None and len(result.energy_input_cum):
                        run.energy_input_total = float(result.energy_input_cum[-1])
                    if result.energy_brake_cum is not None and len(result.energy_brake_cum):
                        run.energy_brake_total = float(result.energy_brake_cum[-1])

                    report_name = f"{prefix}_report.xlsx"
                    report_path = run_reports_dir / report_name
                    export_results_to_excel(result, report_path)
                    run.report_file.name = f"reports/{run_folder_name}/{report_name}"

                    run.save()

                    calculation_model = build_calculation_model(run, brake_objects, result)
                    calculation_model, analysis_snapshot = enrich_with_basic_analysis(calculation_model)

                    CalculationSnapshot.objects.update_or_create(
                        run=run,
                        defaults={
                            "model_version": calculation_model.model_version,
                            "input_snapshot": calculation_model.input_snapshot(),
                            "result_snapshot": calculation_model.result_snapshot(),
                            "analysis_snapshot": analysis_snapshot,
                            "thermal_snapshot": {},
                        },
                    )

                return redirect("run_detail", run_id=run.id)

            except ValueError as exc:
                form.add_error(None, str(exc))
    else:
        initial_main, brakes_initial = _build_initial_from_run(request.GET.get("from_run"))
        form = CalculationForm(initial=initial_main)

        if brakes_initial:
            brake_formset = MagneticBrakeFormSet(initial=brakes_initial, prefix="brakes")
        else:
            brake_formset = MagneticBrakeFormSet(initial=[{}, {}], prefix="brakes")

    runs = CalculationRun.objects.order_by("-created_at")[:20]

    # Срез 3b: каталог тормозов для выбора в форме.
    # Передаём только нужные поля для JS-подстановки (без сложных полей).
    catalog_qs = BrakeCatalog.objects.order_by("name")
    catalog_items = []
    for c in catalog_qs:
        catalog_items.append({
            "id": c.pk,
            "name": c.name,
            "description": c.description or "",
            "model_type": c.model_type,
            "is_parametric": c.is_parametric,
            "is_curve": c.is_curve,
            "summary": c.short_summary,
            # Полный набор параметров — JS подставит в поля формы при выборе
            "params": {
                "gamma": c.gamma,
                "delta": c.delta,
                "n":     c.n,
                "xm":    c.xm,
                "ym":    c.ym,
                "dh1":   c.dh1,
                "dh2":   c.dh2,
                "dm":    c.dm,
                "mu":    c.mu,
                "bz":    c.bz,
                "lya":   c.lya,
                "wn0":   c.wn0,
            },
        })

    return render(
        request,
        "recoil_app/index.html",
        {
            "form": form,
            "brake_formset": brake_formset,
            "runs": runs,
            "catalog_items": catalog_items,
            "catalog_count": len(catalog_items),
        },
    )


def run_detail_view(request, run_id):
    run = get_object_or_404(CalculationRun, pk=run_id)
    brakes = list(run.brakes.order_by("index"))

    chart_fields = [
        "chart_x_t",
        "chart_v_a_t",
        "chart_v_x",
        "chart_fmag_v",
        "chart_forces_secondary",
        "chart_x_t_recoil",
        "chart_v_a_t_recoil",
        "chart_forces_main_recoil",
        "chart_forces_secondary_recoil",
        "chart_x_t_return",
        "chart_v_a_t_return",
        "chart_forces_secondary_return",
    ]

    chart_html: dict[str, str] = {}

    for field_name in chart_fields:
        field_file = getattr(run, field_name, None)
        if field_file:
            try:
                chart_html[field_name] = _read_chart_fragment(field_file.path)
            except FileNotFoundError:
                chart_html[field_name] = ""

    analysis_snapshot = {}
    phase_analysis = {}
    characteristic_points = {}
    engineering_metrics = {}

    try:
        snapshot = run.snapshot
        analysis_snapshot = snapshot.analysis_snapshot or {}
        phase_analysis = analysis_snapshot.get("phase_analysis", {})
        characteristic_points = analysis_snapshot.get("characteristic_points", {})
        engineering_metrics = analysis_snapshot.get("engineering_metrics", {})
    except CalculationSnapshot.DoesNotExist:
        analysis_snapshot = {}
        phase_analysis = {}
        characteristic_points = {}
        engineering_metrics = {}

    return render(
        request,
        "recoil_app/run_detail.html",
        {
            "run": run,
            "brakes": brakes,
            "chart_html": chart_html,
            "analysis_snapshot": analysis_snapshot,
            "phase_analysis": phase_analysis,
            "characteristic_points": characteristic_points,
            "engineering_metrics": engineering_metrics,
        },
    )


def run_detail_v2_view(request, run_id):
    """Новый дизайн страницы результата.

    Отдаёт расчёт с KPI-карточками, аннотированным главным графиком x(t) и энергобалансом.
    Данные те же, что в run_detail_view; меняется только шаблон и состав показываемых графиков.
    """
    run = get_object_or_404(CalculationRun, pk=run_id)
    brakes = list(run.brakes.order_by("index"))

    # Все доступные поля графиков на CalculationRun
    chart_fields = [
        # v2-специфичные
        "chart_x_t_annotated",
        "chart_energy",
        # общие
        "chart_x_t",
        "chart_v_a_t",
        "chart_v_x",
        "chart_fmag_v",
        "chart_forces_secondary",
        # фаза отката
        "chart_x_t_recoil",
        "chart_v_a_t_recoil",
        "chart_forces_main_recoil",
        "chart_forces_secondary_recoil",
        # фаза наката
        "chart_x_t_return",
        "chart_v_a_t_return",
        "chart_forces_secondary_return",
    ]

    chart_html: dict[str, str] = {}
    chart_errors: list[str] = []
    for field_name in chart_fields:
        field_file = getattr(run, field_name, None)
        if field_file and getattr(field_file, "name", ""):
            try:
                chart_html[field_name] = _read_chart_fragment(field_file.path)
            except (FileNotFoundError, OSError) as exc:
                chart_html[field_name] = ""
                chart_errors.append(f"{field_name}: файл не найден ({exc})")
            except UnicodeDecodeError as exc:
                chart_html[field_name] = ""
                chart_errors.append(f"{field_name}: ошибка кодировки ({exc})")
            except Exception as exc:  # noqa: BLE001
                chart_html[field_name] = ""
                chart_errors.append(f"{field_name}: {type(exc).__name__}: {exc}")

    # Если новых графиков нет (старый расчёт) — используем старые как fallback
    has_annotated = bool(chart_html.get("chart_x_t_annotated"))
    has_energy = bool(chart_html.get("chart_energy"))
    has_x_t = bool(chart_html.get("chart_x_t"))

    snapshot_parts = _extract_snapshot_parts(run)

    # KPI-группы для модульной системы скрытия/показа
    kpi_groups = _build_kpi_groups(run, snapshot_parts)

    # Краткая статистика энергобаланса (если есть)
    energy_summary = None
    if run.energy_residual_pct is not None or run.energy_input_total is not None:
        energy_summary = {
            "input_total": run.energy_input_total,
            "brake_total": run.energy_brake_total,
            "residual_pct": run.energy_residual_pct,
        }

    # Какие наборы графиков видимы (для управления модулями)
    has_charts_main = any(chart_html.get(k) for k in [
        "chart_x_t", "chart_v_a_t", "chart_v_x", "chart_fmag_v", "chart_forces_secondary"
    ])
    has_charts_recoil = any(chart_html.get(k) for k in [
        "chart_x_t_recoil", "chart_v_a_t_recoil", "chart_forces_main_recoil", "chart_forces_secondary_recoil"
    ])
    has_charts_return = any(chart_html.get(k) for k in [
        "chart_x_t_return", "chart_v_a_t_return", "chart_forces_secondary_return"
    ])

    return render(
        request,
        "recoil_app/run_detail_v2.html",
        {
            "run": run,
            "brakes": brakes,
            "chart_html": chart_html,
            "has_annotated": has_annotated,
            "has_energy": has_energy,
            "has_x_t": has_x_t,
            "has_charts_main": has_charts_main,
            "has_charts_recoil": has_charts_recoil,
            "has_charts_return": has_charts_return,
            "chart_errors": chart_errors,
            "kpi_groups": kpi_groups,
            "energy_summary": energy_summary,
            "phase_analysis": snapshot_parts["phase_analysis"],
            "characteristic_points": snapshot_parts["characteristic_points"],
            "engineering_metrics": snapshot_parts["engineering_metrics"],
        },
    )


def _build_kpi_groups(run: CalculationRun, snapshot_parts: dict) -> list[dict]:
    """Группированные KPI-карточки для страницы результата.

    Каждая группа: {key, label, cards}
    Каждая карточка: {label, value, unit, hint, status, accent, mono?}
        status: 'ok' | 'warn' | 'danger' | 'neutral'
        accent: 'blue' | 'red' | 'green' | 'amber' | 'purple' | 'gray'
    """
    eng = snapshot_parts.get("engineering_metrics") or {}
    chars = snapshot_parts.get("characteristic_points") or {}
    phase_recoil = (snapshot_parts.get("phase_analysis") or {}).get("recoil") or {}
    phase_return = (snapshot_parts.get("phase_analysis") or {}).get("return") or {}

    groups: list[dict] = []

    # === ГРУППА 1: ОБЗОР ===
    overview: list[dict] = []

    if run.x_max is not None:
        time_hint = ""
        if chars.get("x_max"):
            time_hint = f"t = {chars['x_max'].get('time', 0):.3f} с"
        overview.append({
            "label": "Макс. откат",
            "value": f"{run.x_max * 1000:.1f}",
            "unit": "мм",
            "hint": time_hint,
            "status": "ok",
            "accent": "blue",
        })

    a_abs_max = (eng.get("a_range") or {}).get("abs_max")
    if a_abs_max is not None:
        overview.append({
            "label": "Пик ускор.",
            "value": f"{a_abs_max / 9.81:.1f}",
            "unit": "g",
            "hint": f"{a_abs_max:.0f} м/с²",
            "status": "ok",
            "accent": "amber",
        })

    v_abs_max = (eng.get("v_range") or {}).get("abs_max")
    if v_abs_max is not None:
        time_hint = ""
        if chars.get("v_max"):
            time_hint = f"t = {chars['v_max'].get('time', 0):.3f} с"
        overview.append({
            "label": "Макс. скорость",
            "value": f"{v_abs_max:.2f}",
            "unit": "м/с",
            "hint": time_hint,
            "status": "neutral",
            "accent": "green",
        })

    if run.return_end_time is not None:
        overview.append({
            "label": "Время цикла",
            "value": f"{run.return_end_time:.3f}",
            "unit": "с",
            "hint": "откат + накат",
            "status": "ok",
            "accent": "purple",
        })
    elif run.recoil_end_time is not None:
        overview.append({
            "label": "Время отката",
            "value": f"{run.recoil_end_time:.3f}",
            "unit": "с",
            "hint": "до пика",
            "status": "neutral",
            "accent": "purple",
        })

    if phase_recoil.get("available") and phase_recoil.get("duration") is not None:
        overview.append({
            "label": "Длит. отката",
            "value": f"{phase_recoil['duration']:.3f}",
            "unit": "с",
            "hint": "",
            "status": "neutral",
            "accent": "blue",
        })

    if phase_return.get("available") and phase_return.get("duration") is not None:
        overview.append({
            "label": "Длит. наката",
            "value": f"{phase_return['duration']:.3f}",
            "unit": "с",
            "hint": "",
            "status": "neutral",
            "accent": "blue",
        })

    if overview:
        groups.append({"key": "overview", "label": "Обзор", "cards": overview})

    # === ГРУППА 2: ЭНЕРГОБАЛАНС ===
    energy: list[dict] = []

    if run.energy_input_total is not None:
        energy.append({
            "label": "Подведено",
            "value": f"{run.energy_input_total / 1000:.2f}",
            "unit": "кДж",
            "hint": "выстрел + гравитация",
            "status": "neutral",
            "accent": "amber",
        })

    if run.energy_brake_total is not None:
        ratio_text = ""
        if run.energy_input_total and abs(run.energy_input_total) > 1e-6:
            ratio = run.energy_brake_total / abs(run.energy_input_total) * 100.0
            ratio_text = f"{ratio:.1f}% от входа"
        energy.append({
            "label": "Рассеяно тормозами",
            "value": f"{run.energy_brake_total / 1000:.2f}",
            "unit": "кДж",
            "hint": ratio_text,
            "status": "ok",
            "accent": "red",
        })

    if run.energy_residual_pct is not None:
        if run.energy_residual_pct < 1.0:
            status = "ok"
        elif run.energy_residual_pct < 3.0:
            status = "warn"
        else:
            status = "danger"
        energy.append({
            "label": "Невязка энерг.",
            "value": f"{run.energy_residual_pct:.2f}",
            "unit": "%",
            "hint": "норма < 1%",
            "status": status,
            "accent": "red" if status == "danger" else "amber" if status == "warn" else "green",
        })

    if energy:
        groups.append({"key": "energy", "label": "Энергобаланс", "cards": energy})

    # === ГРУППА 3: ФАЗА ОТКАТА ===
    if phase_recoil.get("available"):
        recoil_cards = [
            _range_card("X_РАЗМАХ", phase_recoil.get("x_min"), phase_recoil.get("x_max"), "м",   "blue"),
            _range_card("V_РАЗМАХ", phase_recoil.get("v_min"), phase_recoil.get("v_max"), "м/с", "green"),
            _range_card("A_РАЗМАХ", phase_recoil.get("a_min"), phase_recoil.get("a_max"), "м/с²","amber"),
        ]
        recoil_cards = [c for c in recoil_cards if c is not None]
        if recoil_cards:
            groups.append({"key": "phase-recoil", "label": "Фаза отката", "cards": recoil_cards})

    # === ГРУППА 4: ФАЗА НАКАТА ===
    if phase_return.get("available"):
        return_cards = [
            _range_card("X_РАЗМАХ", phase_return.get("x_min"), phase_return.get("x_max"), "м",   "blue"),
            _range_card("V_РАЗМАХ", phase_return.get("v_min"), phase_return.get("v_max"), "м/с", "green"),
            _range_card("A_РАЗМАХ", phase_return.get("a_min"), phase_return.get("a_max"), "м/с²","amber"),
        ]
        return_cards = [c for c in return_cards if c is not None]
        if return_cards:
            groups.append({"key": "phase-return", "label": "Фаза наката", "cards": return_cards})

    # === ГРУППА 5: КОНЕЧНОЕ СОСТОЯНИЕ ===
    final: list[dict] = []
    if run.x_final is not None:
        final.append({
            "label": "x конечное",
            "value": _smart_format(run.x_final),
            "unit": "м",
            "hint": "",
            "status": "neutral",
            "accent": "gray",
        })
    if run.v_final is not None:
        final.append({
            "label": "v конечное",
            "value": _smart_format(run.v_final),
            "unit": "м/с",
            "hint": "",
            "status": "neutral",
            "accent": "gray",
        })
    if run.a_final is not None:
        final.append({
            "label": "a конечное",
            "value": _smart_format(run.a_final),
            "unit": "м/с²",
            "hint": "",
            "status": "neutral",
            "accent": "gray",
        })

    if final:
        groups.append({"key": "final-state", "label": "Конечное состояние", "cards": final})

    return groups


def _range_card(label: str, vmin, vmax, unit: str, accent: str) -> dict | None:
    """Карточка с диапазоном min..max (для парных значений X / V / A в фазах).

    Структура:
        label = "X_РАЗМАХ"
        value = main число (max — оно обычно «главное»)
        subvalues = [{label: "min", value: ...}, {label: "max", value: ...}]

    None — если оба значения отсутствуют.
    """
    if vmin is None and vmax is None:
        return None
    return {
        "label": label,
        "value": _smart_format(vmax if vmax is not None else vmin),
        "unit": unit,
        "hint": "",
        "status": "neutral",
        "accent": accent,
        "subvalues": [
            {"label": "min", "value": _smart_format(vmin) if vmin is not None else "—", "unit": unit},
            {"label": "max", "value": _smart_format(vmax) if vmax is not None else "—", "unit": unit},
        ],
    }


def _smart_format(value) -> str:
    """Умное форматирование числа: маленькие — научной записью, остальные — фикс. знаков."""
    if value is None:
        return "—"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    av = abs(v)
    if av == 0.0:
        return "0"
    if av < 0.001:
        return f"{v:.2e}"
    if av < 1.0:
        return f"{v:.4f}"
    if av < 100.0:
        return f"{v:.3f}"
    if av < 10000.0:
        return f"{v:.1f}"
    return f"{v:.0f}"


@require_POST
def delete_run_view(request, run_id):
    run = get_object_or_404(CalculationRun, pk=run_id)

    folder_path: Path | None = None
    if run.report_file and run.report_file.name:
        folder_path = Path(settings.MEDIA_ROOT) / Path(run.report_file.name).parent

    for brake in run.brakes.all():
        if brake.curve_file:
            try:
                brake.curve_file.delete(save=False)
            except Exception:
                pass

    file_fields = [
        "input_file",
        "report_file",
        "chart_x_t",
        "chart_v_a_t",
        "chart_v_x",
        "chart_fmag_v",
        "chart_forces_secondary",
        "chart_x_t_recoil",
        "chart_v_a_t_recoil",
        "chart_forces_main_recoil",
        "chart_forces_secondary_recoil",
        "chart_x_t_return",
        "chart_v_a_t_return",
        "chart_forces_secondary_return",
        # v2:
        "chart_x_t_annotated",
        "chart_energy",
    ]

    for field_name in file_fields:
        field_file = getattr(run, field_name, None)
        if field_file:
            try:
                field_file.delete(save=False)
            except Exception:
                pass

    run.delete()

    if folder_path and folder_path.exists():
        shutil.rmtree(folder_path, ignore_errors=True)

    messages.success(request, "Расчёт и его файлы удалены.")
    return redirect("dashboard")


# ============================================================================
# СРЕЗ 2: ДАШБОРД
# ============================================================================

def dashboard_view(request):
    """Главный экран приложения — список расчётов с фильтрами и stat-карточками.

    Query-параметры:
        q       — текстовый поиск по имени расчёта
        filter  — 'all' | 'success' | 'warnings' | 'recent'
        sort    — '-created_at' | 'created_at' | 'name' | '-x_max' | 'return_end_time'
        page    — номер страницы (пагинация по 20 на страницу)
    """
    # ---- 1. Stat-карточки (считаются по полной БД, без фильтров) ----
    week_ago = timezone.now() - timedelta(days=7)

    stats = {
        "total":     CalculationRun.objects.count(),
        "last_week": CalculationRun.objects.filter(created_at__gte=week_ago).count(),
        "success":   CalculationRun.objects.filter(termination_reason="returned_to_zero").count(),
        "warnings":  CalculationRun.objects.filter(
            Q(spring_out_of_range=True) | Q(warnings_text__gt="")
        ).count(),
        "catalog":   BrakeCatalog.objects.count(),
    }

    # ---- 2. Фильтрация и поиск ----
    qs = CalculationRun.objects.all().prefetch_related("brakes")

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(name__icontains=q)

    flt = request.GET.get("filter") or "all"
    if flt == "success":
        qs = qs.filter(termination_reason="returned_to_zero")
    elif flt == "warnings":
        qs = qs.filter(Q(spring_out_of_range=True) | Q(warnings_text__gt=""))
    elif flt == "recent":
        qs = qs.filter(created_at__gte=week_ago)

    # ---- 3. Сортировка ----
    sort = request.GET.get("sort") or "-created_at"
    allowed_sorts = {
        "-created_at":      "-created_at",
        "created_at":       "created_at",
        "name":             "name",
        "-name":            "-name",
        "-x_max":           "-x_max",
        "return_end_time":  "return_end_time",
    }
    sort_field = allowed_sorts.get(sort, "-created_at")
    # null'ы в конец при сортировках по числам (модели не всегда заполнены)
    qs = qs.order_by(sort_field, "-created_at")

    # ---- 4. Пагинация ----
    paginator = Paginator(qs, 20)
    page_num = request.GET.get("page") or 1
    page = paginator.get_page(page_num)

    # ---- 5. Карточки расчётов: вытаскиваем агрегированную инфу для каждого ----
    cards = []
    for run in page.object_list:
        # Определяем визуальный статус для бейджа
        if run.termination_reason == "returned_to_zero":
            status = "ok"
            status_label = "завершён"
        elif run.termination_reason == "time_limit":
            status = "warn"
            status_label = "по времени"
        elif run.termination_reason:
            status = "warn"
            status_label = run.termination_reason
        else:
            status = "neutral"
            status_label = "—"

        has_warnings = bool(run.spring_out_of_range or (run.warnings_text or "").strip())

        cards.append({
            "run": run,
            "status": status,
            "status_label": status_label,
            "has_warnings": has_warnings,
        })

    # ---- 6. Подготовка query-string для пагинации (сохранить фильтры между страницами) ----
    qs_keep: list[str] = []
    for key in ("q", "filter", "sort"):
        val = request.GET.get(key)
        if val:
            qs_keep.append(f"{key}={val}")
    qs_keep_str = "&".join(qs_keep)

    return render(
        request,
        "recoil_app/dashboard.html",
        {
            "stats": stats,
            "cards": cards,
            "page": page,
            "paginator": paginator,
            "q": q,
            "filter_value": flt,
            "sort_value": sort,
            "qs_keep_str": qs_keep_str,
        },
    )

# ============================================================================
# СРЕЗ 3a: Views для каталога тормозов
# ============================================================================

def catalog_list_view(request):
    """Список всех тормозов в каталоге с поиском и сортировкой."""
    qs = BrakeCatalog.objects.all()

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))

    flt = request.GET.get("filter") or "all"
    if flt == "parametric":
        qs = qs.filter(model_type=BrakeCatalog.MODEL_TYPE_PARAMETRIC)
    elif flt == "curve":
        qs = qs.filter(model_type=BrakeCatalog.MODEL_TYPE_CURVE)

    sort = request.GET.get("sort") or "name"
    allowed = {
        "name": "name",
        "-name": "-name",
        "-created_at": "-created_at",
        "created_at": "created_at",
        "-updated_at": "-updated_at",
    }
    qs = qs.order_by(allowed.get(sort, "name"))

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get("page") or 1)

    qs_keep_parts: list[str] = []
    for key in ("q", "filter", "sort"):
        v = request.GET.get(key)
        if v:
            qs_keep_parts.append(f"{key}={v}")
    qs_keep_str = "&".join(qs_keep_parts)

    return render(
        request,
        "recoil_app/catalog_list.html",
        {
            "page": page,
            "paginator": paginator,
            "items": page.object_list,
            "q": q,
            "filter_value": flt,
            "sort_value": sort,
            "qs_keep_str": qs_keep_str,
            "total_count": BrakeCatalog.objects.count(),
            "parametric_count": BrakeCatalog.objects.filter(model_type=BrakeCatalog.MODEL_TYPE_PARAMETRIC).count(),
            "curve_count": BrakeCatalog.objects.filter(model_type=BrakeCatalog.MODEL_TYPE_CURVE).count(),
        },
    )


def catalog_new_view(request):
    """Создание нового тормоза в каталоге."""
    if request.method == "POST":
        form = BrakeCatalogForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f"Тормоз «{obj.name}» добавлен в каталог.")
            return redirect("catalog_list")
    else:
        form = BrakeCatalogForm()

    return render(
        request,
        "recoil_app/catalog_form.html",
        {
            "form": form,
            "is_edit": False,
            "object": None,
        },
    )


def catalog_detail_view(request, pk):
    """Детальная страница тормоза в каталоге.

    Показывает:
    - Hero с именем, описанием, бейджем типа
    - Для параметрических: KPI-карточки с физическими параметрами + формула
    - Для табличных: график F(v), статистика по точкам, таблица первых N точек
    - Кнопки: редактировать, удалить, назад к списку
    """
    from .services.charting import make_brake_curve_fragment

    obj = get_object_or_404(BrakeCatalog, pk=pk)

    # === Параметрические карточки ===
    parametric_cards: list[dict] = []
    if obj.is_parametric:
        # Группируем параметры на «материал/электрика» и «геометрия»
        param_specs = [
            # (model_attr, label, symbol, unit, accent, group)
            ("gamma", "Удельная проводимость",        "γ",            "(Ом·м)⁻¹", "blue",   "material"),
            ("delta", "Толщина шины",                  "δ",            "м",     "blue",   "material"),
            ("mu",    "Магнитная проницаемость шины",  "μ",            "Гн/м",  "blue",   "material"),
            ("bz",    "Индукция в рабочем зазоре",     "B̄₃",           "Тл",    "amber",  "material"),
            ("n",     "Количество блоков",             "N",            "",      "purple", "geometry"),
            ("xm",    "Размер магнита по оси X",       "x_m",          "м",     "green",  "geometry"),
            ("ym",    "Размер магнита по оси Y",       "y_m",          "м",     "green",  "geometry"),
            ("dh1",   "Выступ 1-го края шины",         "Δh₁",          "м",     "green",  "geometry"),
            ("dh2",   "Выступ 2-го края шины",         "Δh₂",          "м",     "green",  "geometry"),
            ("dm",    "Промежутки между магнитами",    "d_m",          "м",     "green",  "geometry"),
            ("lya",   "Параметр λa",                   "λa",           "",      "red",    "extra"),
            ("wn0",   "Начальное состояние wn",        "w_n0",         "",      "red",    "extra"),
        ]

        for attr, label, symbol, unit, accent, group in param_specs:
            value = getattr(obj, attr)
            parametric_cards.append({
                "label":  label,
                "symbol": symbol,
                "value":  value,
                "unit":   unit,
                "accent": accent,
                "group":  group,
            })

    # === Точки кривой F(v) и Plotly-фрагмент ===
    curve_html = ""
    curve_points: list[dict] = []
    curve_stats = None
    curve_error = None

    if obj.is_curve and obj.curve_file:
        try:
            obj.curve_file.open("rb")
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(obj.curve_file, read_only=True, data_only=True)
                try:
                    helper_form = MagneticBrakeForm()
                    curve_points = helper_form._parse_force_curve_sheet(workbook.active)
                finally:
                    workbook.close()
            finally:
                obj.curve_file.close()

            if curve_points:
                curve_html = make_brake_curve_fragment(
                    curve_points,
                    title=f"F(v) — {obj.name}",
                )
                vs = [p["velocity"] for p in curve_points]
                fs = [p["force"] for p in curve_points]
                curve_stats = {
                    "n_points": len(curve_points),
                    "v_min": min(vs),
                    "v_max": max(vs),
                    "f_min": min(fs),
                    "f_max": max(fs),
                }
        except Exception as exc:  # noqa: BLE001
            curve_error = f"Не удалось прочитать файл F(v): {type(exc).__name__}: {exc}"

    # === Использование тормоза в расчётах ===
    # Тормоз в каталоге не связан напрямую с MagneticBrakeConfig (copy-on-use),
    # но мы можем показать сколько раз он был использован — по совпадению имени.
    # Это эвристика, не точный учёт; для презентации достаточно.
    usage_count = MagneticBrakeConfig.objects.filter(name=obj.name).count()

    return render(
        request,
        "recoil_app/catalog_detail.html",
        {
            "object": obj,
            "parametric_cards": parametric_cards,
            "curve_points": curve_points,
            "curve_html": curve_html,
            "curve_stats": curve_stats,
            "curve_error": curve_error,
            "usage_count": usage_count,
        },
    )


def catalog_edit_view(request, pk):
    """Редактирование существующей записи каталога."""
    obj = get_object_or_404(BrakeCatalog, pk=pk)

    if request.method == "POST":
        form = BrakeCatalogForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"Тормоз «{obj.name}» обновлён.")
            return redirect("catalog_list")
    else:
        form = BrakeCatalogForm(instance=obj)

    return render(
        request,
        "recoil_app/catalog_form.html",
        {
            "form": form,
            "is_edit": True,
            "object": obj,
        },
    )


@require_POST
def catalog_delete_view(request, pk):
    """Удаление тормоза из каталога."""
    obj = get_object_or_404(BrakeCatalog, pk=pk)
    name = obj.name
    # Удалим файл кривой, если есть
    if obj.curve_file:
        try:
            obj.curve_file.delete(save=False)
        except Exception:
            pass
    obj.delete()
    messages.success(request, f"Тормоз «{name}» удалён из каталога.")
    return redirect("catalog_list")


@require_POST
def catalog_save_from_brake_form_view(request):
    """AJAX-эндпоинт для сохранения тормоза из формы расчёта в каталог.

    Принимает JSON-payload или form-encoded данные с параметрами тормоза.
    Возвращает JSON: {"ok": true, "id": N, "name": "..."} или {"ok": false, "error": "..."}.
    """
    from django.http import JsonResponse

    def _f(key):
        """Получить float или None из POST."""
        v = (request.POST.get(key) or "").strip()
        if not v:
            return None
        try:
            return float(v.replace(",", "."))
        except (ValueError, TypeError):
            return None

    def _i(key):
        v = _f(key)
        if v is None:
            return None
        try:
            return int(v)
        except (ValueError, TypeError):
            return None

    name = (request.POST.get("name") or "").strip()
    description = (request.POST.get("description") or "").strip()
    model_type = (request.POST.get("model_type") or "").strip()

    if not name:
        return JsonResponse({"ok": False, "error": "Имя тормоза обязательно."}, status=400)

    if model_type not in (BrakeCatalog.MODEL_TYPE_PARAMETRIC, BrakeCatalog.MODEL_TYPE_CURVE):
        return JsonResponse({"ok": False, "error": "Некорректный тип модели."}, status=400)

    if BrakeCatalog.objects.filter(name=name).exists():
        return JsonResponse(
            {"ok": False, "error": f"Тормоз с именем «{name}» уже есть в каталоге."},
            status=400,
        )

    if model_type == BrakeCatalog.MODEL_TYPE_PARAMETRIC:
        # Все параметры обязательные
        params = {
            "gamma": _f("gamma"),
            "delta": _f("delta"),
            "n":     _i("n"),
            "xm":    _f("xm"),
            "ym":    _f("ym"),
            "dh1":   _f("dh1"),
            "dh2":   _f("dh2"),
            "dm":    _f("dm"),
            "mu":    _f("mu"),
            "bz":    _f("bz"),
            "lya":   _f("lya"),
            "wn0":   _f("wn0"),
        }
        missing = [k for k, v in params.items() if v is None]
        if missing:
            return JsonResponse(
                {"ok": False, "error": f"Не заполнены параметры: {', '.join(missing)}"},
                status=400,
            )
        obj = BrakeCatalog.objects.create(
            name=name,
            description=description,
            model_type=model_type,
            **params,
        )
    else:
        # Для табличной — нужен файл F(v).
        # При сохранении из формы расчёта файл может быть только что загружен в форму
        # (в multipart-запросе) — либо его уже нет (файл из catalog_source_id).
        curve_file = request.FILES.get("curve_file")
        catalog_source_id = (request.POST.get("catalog_source_id") or "").strip()

        if not curve_file and not catalog_source_id:
            return JsonResponse(
                {"ok": False, "error": "Для табличной модели нужен файл F(v) или ссылка на каталог."},
                status=400,
            )

        if curve_file:
            obj = BrakeCatalog.objects.create(
                name=name,
                description=description,
                model_type=model_type,
                curve_file=curve_file,
            )
        else:
            # Копируем файл из родительской записи каталога
            try:
                source = BrakeCatalog.objects.get(pk=int(catalog_source_id))
            except (ValueError, BrakeCatalog.DoesNotExist):
                return JsonResponse(
                    {"ok": False, "error": "Источник в каталоге не найден."},
                    status=400,
                )
            obj = BrakeCatalog.objects.create(
                name=name,
                description=description,
                model_type=model_type,
            )
            if source.curve_file:
                source.curve_file.open("rb")
                try:
                    content = source.curve_file.read()
                finally:
                    source.curve_file.close()
                original_name = Path(source.curve_file.name).name or "curve.xlsx"
                obj.curve_file.save(original_name, ContentFile(content), save=True)

    return JsonResponse({
        "ok": True,
        "id": obj.pk,
        "name": obj.name,
        "model_type": obj.model_type,
    })
