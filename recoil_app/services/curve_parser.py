"""Парсер табличной характеристики F(v) тормоза из Excel-файла.

Отдельный сервис, чтобы forms и views могли использовать одинаковую логику
без создания фиктивного `MagneticBrakeForm()` ради метода.

Все ошибки выбрасываются как `django.core.exceptions.ValidationError` —
это совместимо с использованием в `Form.clean()` и легко перехватывается
во views через `except ValidationError` или `except Exception`.
"""

from __future__ import annotations

from pathlib import Path

from django.core.exceptions import ValidationError
from openpyxl import load_workbook


_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def parse_force_curve_file(uploaded_file) -> list[dict]:
    """Прочитать Excel-файл и вернуть список точек {order, velocity, force}.

    Парсится первый лист (`workbook.active`).
    """
    _validate_excel_extension(uploaded_file.name)

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(
            "Не удалось прочитать Excel-файл характеристики тормоза."
        ) from exc
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    try:
        return parse_force_curve_sheet(workbook.active)
    finally:
        workbook.close()


def parse_force_curve_sheet(sheet) -> list[dict]:
    """Распарсить лист openpyxl в список точек F(v).

    Формат: колонка A — скорость, колонка B — сила. Первая строка может быть
    заголовком (если первое значение нечисловое — пропускается).

    Валидация:
    - минимум 2 точки
    - скорости и силы ≥ 0
    - скорости строго возрастающие
    """
    parsed_points: list[dict] = []
    header_skipped = False

    for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        velocity_raw = row[0] if len(row) > 0 else None
        force_raw = row[1] if len(row) > 1 else None

        if _is_empty_cell(velocity_raw) and _is_empty_cell(force_raw):
            continue

        if _is_empty_cell(velocity_raw) or _is_empty_cell(force_raw):
            raise ValidationError(
                f"Строка {row_no}: должны быть заполнены и скорость, и сила."
            )

        velocity = _coerce_excel_number(velocity_raw)
        force = _coerce_excel_number(force_raw)

        # Допускаем одну строку заголовков в начале файла
        if velocity is None or force is None:
            if not header_skipped and not parsed_points:
                header_skipped = True
                continue
            raise ValidationError(
                f"Строка {row_no}: значения скорости и силы должны быть числами."
            )

        if velocity < 0:
            raise ValidationError(
                f"Строка {row_no}: скорость не может быть отрицательной."
            )
        if force < 0:
            raise ValidationError(
                f"Строка {row_no}: сила торможения не может быть отрицательной."
            )

        parsed_points.append(
            {
                "order": len(parsed_points) + 1,
                "velocity": velocity,
                "force": force,
            }
        )

    if len(parsed_points) < 2:
        raise ValidationError("Для графического тормоза необходимо минимум 2 точки.")

    velocities = [point["velocity"] for point in parsed_points]
    if len(set(velocities)) != len(velocities):
        raise ValidationError("Скорости в характеристике тормоза не должны повторяться.")
    if velocities != sorted(velocities):
        raise ValidationError(
            "Скорости в характеристике тормоза должны быть строго возрастающими."
        )

    return parsed_points


def _validate_excel_extension(filename: str) -> None:
    suffix = Path(filename).suffix.lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise ValidationError(
            "Поддерживаются только Excel-файлы форматов .xlsx, .xlsm, .xltx, .xltm."
        )


def _is_empty_cell(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _coerce_excel_number(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(",", ".")
        if not normalized:
            return None
        try:
            return float(normalized)
        except ValueError:
            return None
    return None
