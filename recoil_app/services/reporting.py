from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font


def _autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def _write_sheet_from_table(wb: Workbook, title: str, headers: list[str], rows: list[list[float | int | str]]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    _autosize_columns(ws)


def _add_scatter_chart_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[float | int | str]],
    x_col: int,
    y_cols: list[int],
    chart_title: str,
    x_title: str,
    y_title: str,
) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    chart = ScatterChart()
    chart.title = chart_title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.style = 2
    chart.height = 16
    chart.width = 28

    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=ws.max_row)
    for y_col in y_cols:
        values = Reference(ws, min_col=y_col, min_row=1, max_row=ws.max_row)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

    ws.add_chart(chart, "H2")
    _autosize_columns(ws)


def export_results_to_excel(result, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"

    ws_summary.append(["metric", "value"])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    recoil_end_index = result.recoil_end_index if result.recoil_end_index is not None else -1
    return_end_index = result.return_end_index if result.return_end_index is not None else -1

    ws_summary.append(["x_max", float(result.x.max())])
    ws_summary.append(["v_max", float(result.v.max())])
    ws_summary.append(["x_final", float(result.x[-1])])
    ws_summary.append(["v_final", float(result.v[-1])])
    ws_summary.append(["a_final", float(result.a[-1])])
    ws_summary.append(["recoil_end_time", result.recoil_end_time])
    ws_summary.append(["return_end_time", result.return_end_time])
    ws_summary.append(["recoil_end_index", recoil_end_index])
    ws_summary.append(["return_end_index", return_end_index])
    ws_summary.append(["termination_reason", result.termination_reason])
    ws_summary.append(["spring_out_of_range", result.spring_out_of_range])

    if result.warnings:
        ws_summary.append(["warnings", " | ".join(result.warnings)])

    _autosize_columns(ws_summary)

    n_brakes = result.f_magnetic_each.shape[1] if result.f_magnetic_each.ndim == 2 else 0

    headers = [
        "t, c",
        "x, м",
        "v, м/с",
        "a, м/с²",
        "Fобщ, Н",
        "Fдв, Н",
        "Fпруж, Н",
        "Fмаг_сумм, Н",
        "Fугла, Н",
    ] + [f"Fмаг{j + 1}, Н" for j in range(n_brakes)]

    rows = []
    for i in range(len(result.t)):
        row = [
            float(result.t[i]),
            float(result.x[i]),
            float(result.v[i]),
            float(result.a[i]),
            float(result.f_total[i]),
            float(result.f_ext[i]),
            float(result.f_spring[i]),
            float(result.f_magnetic[i]),
            float(result.f_angle[i]),
        ]
        for j in range(n_brakes):
            row.append(float(result.f_magnetic_each[i, j]))
        rows.append(row)

    _write_sheet_from_table(wb, "data", headers, rows)

    recoil_rows = []
    if result.recoil_end_index is not None:
        for i in range(result.recoil_end_index + 1):
            row = [
                float(result.t[i]),
                float(result.x[i]),
                float(result.v[i]),
                float(result.a[i]),
                float(result.f_total[i]),
                float(result.f_ext[i]),
                float(result.f_spring[i]),
                float(result.f_magnetic[i]),
                float(result.f_angle[i]),
            ]
            for j in range(n_brakes):
                row.append(float(result.f_magnetic_each[i, j]))
            recoil_rows.append(row)

    return_rows = []
    if result.recoil_end_index is not None:
        for i in range(result.recoil_end_index, len(result.t)):
            row = [
                float(result.t[i]),
                float(result.x[i]),
                float(result.v[i]),
                float(result.a[i]),
                float(result.f_total[i]),
                float(result.f_ext[i]),
                float(result.f_spring[i]),
                float(result.f_magnetic[i]),
                float(result.f_angle[i]),
            ]
            for j in range(n_brakes):
                row.append(float(result.f_magnetic_each[i, j]))
            return_rows.append(row)

    if rows:
        _add_scatter_chart_sheet(
            wb,
            "chart_x_t",
            ["t, c", "x, м"],
            [[r[0], r[1]] for r in rows],
            x_col=1,
            y_cols=[2],
            chart_title="Перемещение от времени",
            x_title="t, c",
            y_title="x, м",
        )

        _add_scatter_chart_sheet(
            wb,
            "chart_v_a_t",
            ["t, c", "v, м/с", "a, м/с²"],
            [[r[0], r[2], r[3]] for r in rows],
            x_col=1,
            y_cols=[2, 3],
            chart_title="Скорость и ускорение от времени",
            x_title="t, c",
            y_title="v / a",
        )

        _add_scatter_chart_sheet(
            wb,
            "chart_v_x",
            ["x, м", "v, м/с"],
            [[r[1], r[2]] for r in rows],
            x_col=1,
            y_cols=[2],
            chart_title="Скорость от перемещения",
            x_title="x, м",
            y_title="v, м/с",
        )

        fmag_headers = ["v, м/с", "Fмаг_сумм, Н"] + [f"Fмаг{j + 1}, Н" for j in range(n_brakes)]
        fmag_rows = []
        for i in range(len(result.t)):
            row = [float(result.v[i]), float(result.f_magnetic[i])]
            for j in range(n_brakes):
                row.append(float(result.f_magnetic_each[i, j]))
            fmag_rows.append(row)

        _add_scatter_chart_sheet(
            wb,
            "chart_fmag_v",
            fmag_headers,
            fmag_rows,
            x_col=1,
            y_cols=list(range(2, 2 + len(fmag_headers) - 1)),
            chart_title="Магнитные силы от скорости",
            x_title="v, м/с",
            y_title="F, Н",
        )

        secondary_headers = ["t, c", "Fугла, Н", "Fпруж, Н", "Fмаг_сумм, Н"] + [f"Fмаг{j + 1}, Н" for j in range(n_brakes)]
        secondary_rows = []
        for i in range(len(result.t)):
            row = [
                float(result.t[i]),
                float(result.f_angle[i]),
                float(result.f_spring[i]),
                float(result.f_magnetic[i]),
            ]
            for j in range(n_brakes):
                row.append(float(result.f_magnetic_each[i, j]))
            secondary_rows.append(row)

        _add_scatter_chart_sheet(
            wb,
            "chart_forces_secondary",
            secondary_headers,
            secondary_rows,
            x_col=1,
            y_cols=list(range(2, 2 + len(secondary_headers) - 1)),
            chart_title="Остальные силы от времени",
            x_title="t, c",
            y_title="F, Н",
        )

    if recoil_rows:
        _add_scatter_chart_sheet(
            wb,
            "chart_x_t_recoil",
            ["t, c", "x, м"],
            [[r[0], r[1]] for r in recoil_rows],
            x_col=1,
            y_cols=[2],
            chart_title="Перемещение от времени — откат",
            x_title="t, c",
            y_title="x, м",
        )

        _add_scatter_chart_sheet(
            wb,
            "chart_v_a_t_recoil",
            ["t, c", "v, м/с", "a, м/с²"],
            [[r[0], r[2], r[3]] for r in recoil_rows],
            x_col=1,
            y_cols=[2, 3],
            chart_title="Скорость и ускорение — откат",
            x_title="t, c",
            y_title="v / a",
        )

        _add_scatter_chart_sheet(
            wb,
            "chart_forces_main_recoil",
            ["t, c", "Fдв, Н", "Fобщ, Н"],
            [[r[0], r[5], r[4]] for r in recoil_rows],
            x_col=1,
            y_cols=[2, 3],
            chart_title="Движущая и суммарная силы — откат",
            x_title="t, c",
            y_title="F, Н",
        )

        recoil_secondary_headers = ["t, c", "Fугла, Н", "Fпруж, Н", "Fмаг_сумм, Н"] + [f"Fмаг{j + 1}, Н" for j in range(n_brakes)]
        recoil_secondary_rows = []
        for r in recoil_rows:
            row = [r[0], r[8], r[6], r[7]]
            for j in range(n_brakes):
                row.append(r[9 + j])
            recoil_secondary_rows.append(row)

        _add_scatter_chart_sheet(
            wb,
            "chart_forces_secondary_recoil",
            recoil_secondary_headers,
            recoil_secondary_rows,
            x_col=1,
            y_cols=list(range(2, 2 + len(recoil_secondary_headers) - 1)),
            chart_title="Остальные силы — откат",
            x_title="t, c",
            y_title="F, Н",
        )

    if return_rows:
        _add_scatter_chart_sheet(
            wb,
            "chart_x_t_return",
            ["t, c", "x, м"],
            [[r[0], r[1]] for r in return_rows],
            x_col=1,
            y_cols=[2],
            chart_title="Перемещение от времени — накат",
            x_title="t, c",
            y_title="x, м",
        )

        _add_scatter_chart_sheet(
            wb,
            "chart_v_a_t_return",
            ["t, c", "v, м/с", "a, м/с²"],
            [[r[0], r[2], r[3]] for r in return_rows],
            x_col=1,
            y_cols=[2, 3],
            chart_title="Скорость и ускорение — накат",
            x_title="t, c",
            y_title="v / a",
        )

        return_secondary_headers = ["t, c", "Fугла, Н", "Fпруж, Н", "Fмаг_сумм, Н", "Fобщ, Н"] + [f"Fмаг{j + 1}, Н" for j in range(n_brakes)]
        return_secondary_rows = []
        for r in return_rows:
            row = [r[0], r[8], r[6], r[7], r[4]]
            for j in range(n_brakes):
                row.append(r[9 + j])
            return_secondary_rows.append(row)

        _add_scatter_chart_sheet(
            wb,
            "chart_forces_secondary_return",
            return_secondary_headers,
            return_secondary_rows,
            x_col=1,
            y_cols=list(range(2, 2 + len(return_secondary_headers) - 1)),
            chart_title="Остальные силы — накат",
            x_title="t, c",
            y_title="F, Н",
        )

    wb.save(output_path)
    return output_path