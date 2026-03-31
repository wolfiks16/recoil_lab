from pathlib import Path
import numpy as np
import openpyxl
from .interpolation import LinearTailPchip, prepare_monotonic_nodes

def _to_float(value: object) -> float:
    if value is None:
        raise ValueError("Пустое значение в таблице характеристики.")
    if isinstance(value, str):
        value = value.replace(",", ".").strip()
    return float(value)

def load_recoil_characteristics(xlsx_path: str | Path):
    """
    Ожидаются листы:
    - 'сила от времени'      : столбцы t, F(кН)
    - 'сила от перемещения'  : столбцы X, F(кН)

    Возвращает:
    - F_ext(t) -> Н
    - F_spring(x) -> Н
    - t_ext_max
    - t_range = (t_min, t_max)
    - x_range = (x_min, x_max)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws_t = wb["сила от времени"]
    t_vals = []
    f_drive_vals = []
    for row in ws_t.iter_rows(values_only=True):
        if len(row) >= 3 and row[1] not in (None, "t, с") and row[2] is not None:
            try:
                t_vals.append(_to_float(row[1]))
                f_drive_vals.append(_to_float(row[2]) * 1_000.0)
            except (ValueError, TypeError):
                continue

    ws_x = wb["сила от перемещения"]
    x_vals = []
    f_spring_vals = []
    for row in ws_x.iter_rows(values_only=True):
        if len(row) >= 5 and row[3] is not None and row[4] is not None and row[3] != "X,м":
            try:
                x_vals.append(_to_float(row[3]))
                f_spring_vals.append(abs(_to_float(row[4])) * 1_000.0)
            except (ValueError, TypeError):
                continue

    if not t_vals or not x_vals:
        raise ValueError("Не удалось прочитать характеристики из Excel.")

    t_arr, f_drive_arr = prepare_monotonic_nodes(np.array(t_vals), np.array(f_drive_vals))
    x_arr, f_spring_arr = prepare_monotonic_nodes(np.array(x_vals), np.array(f_spring_vals))

    drive_interp = LinearTailPchip(t_arr, f_drive_arr)
    spring_interp = LinearTailPchip(x_arr, f_spring_arr)

    t_ext_max = float(t_arr[-1])
    x_range = (float(x_arr[0]), float(x_arr[-1]))

    return drive_interp, spring_interp, t_ext_max, x_range