from __future__ import annotations

import re
from pathlib import Path

from django import forms
from django.core.exceptions import ValidationError
from django.core.validators import MaxValueValidator, MinValueValidator
from django.forms import BaseFormSet, formset_factory
from openpyxl import load_workbook

from .models import CalculationRun, MagneticBrakeConfig


class CalculationForm(forms.Form):
    name = forms.CharField(
        required=True,
        label="Название расчёта",
        widget=forms.TextInput(attrs={
            "pattern": "[A-Za-z0-9_-]+",
            "title": "Только английские буквы, цифры, дефис и подчёркивание",
        }),
    )
    input_file = forms.FileField(label="Файл характеристик Excel")

    mass = forms.FloatField(
        label="Масса",
        validators=[MinValueValidator(1e-6, "Масса должна быть положительной.")],
        widget=forms.NumberInput(attrs={"min": "0.000001", "step": "any"}),
    )
    angle_deg = forms.FloatField(
        initial=70.0,
        label="Угол, град",
        validators=[
            MinValueValidator(0.0, "Угол должен быть неотрицательным."),
            MaxValueValidator(90.0, "Угол не должен превышать 90°."),
        ],
        widget=forms.NumberInput(attrs={"min": "0", "max": "90", "step": "any"}),
    )
    v0 = forms.FloatField(
        initial=0.0,
        label="Начальная скорость",
        validators=[MinValueValidator(0.0, "Начальная скорость должна быть ≥ 0.")],
        widget=forms.NumberInput(attrs={"min": "0", "step": "any"}),
    )
    x0 = forms.FloatField(
        initial=0.0,
        label="Начальное перемещение",
        validators=[MinValueValidator(0.0, "Начальное перемещение должно быть ≥ 0.")],
        widget=forms.NumberInput(attrs={"min": "0", "step": "any"}),
    )
    t_max = forms.FloatField(
        initial=0.15,
        label="Время расчёта",
        validators=[
            MinValueValidator(1e-6, "Время расчёта должно быть положительным."),
            MaxValueValidator(10.0, "Время расчёта не должно превышать 10 с."),
        ],
        widget=forms.NumberInput(attrs={"min": "0.000001", "max": "10", "step": "any"}),
    )
    dt = forms.FloatField(initial=1e-4, label="Шаг dt")

    def clean_name(self):
        name = self.cleaned_data["name"].strip()

        if not re.fullmatch(r"[A-Za-z0-9_-]+", name):
            raise forms.ValidationError(
                "Название должно содержать только английские буквы, цифры, дефис и подчёркивание."
            )

        if CalculationRun.objects.filter(name=name).exists():
            raise forms.ValidationError("Расчёт с таким названием уже существует.")

        return name


class MagneticBrakeForm(forms.Form):
    model_type = forms.ChoiceField(
        label="Тип задания тормоза",
        choices=MagneticBrakeConfig.MODEL_TYPE_CHOICES,
        initial=MagneticBrakeConfig.MODEL_TYPE_PARAMETRIC,
    )
    name = forms.CharField(label="Имя тормоза", max_length=255, required=False)

    # Параметрическая модель
    gamma = forms.FloatField(label="gamma", required=False)
    delta = forms.FloatField(label="delta", required=False)
    xm = forms.FloatField(label="xm", required=False)
    ym = forms.FloatField(label="ym", required=False)
    dh1 = forms.FloatField(label="dh1", required=False)
    dh2 = forms.FloatField(label="dh2", required=False)
    dm = forms.FloatField(label="dm", required=False)
    n = forms.IntegerField(
        label="n",
        required=False,
        widget=forms.NumberInput(attrs={"min": "1", "step": "1"}),
    )
    mu = forms.FloatField(label="mu", required=False)
    bz = forms.FloatField(label="bz", required=False)
    lya = forms.FloatField(initial=2.5, label="lya", required=False)
    wn0 = forms.FloatField(initial=1.0, label="wn0", required=False)

    # Табличная модель через Excel
    force_curve_file = forms.FileField(
        label="Excel-файл характеристики F(v)",
        required=False,
        help_text=(
            "Первый лист: колонка A — скорость, колонка B — сила. "
            "Первая строка может быть заголовком."
        ),
    )

    # Для сценария 'Подставить в новый расчёт' — reuse уже сохранённой curve-характеристики
    curve_source_brake_id = forms.CharField(
        required=False,
        widget=forms.HiddenInput(),
    )

    # Срез 3b: ID выбранного тормоза из каталога (если пользователь выбрал «Из каталога»).
    # Параметры всё равно копируются из формы (юзер мог их откорректировать),
    # но для curve — copy-on-use файла F(v) из каталога идёт через этот ID в view'хе.
    catalog_source_id = forms.IntegerField(
        required=False,
        widget=forms.HiddenInput(),
    )

    def clean(self):
        cleaned_data = super().clean()
        model_type = cleaned_data.get("model_type")

        if model_type == MagneticBrakeConfig.MODEL_TYPE_PARAMETRIC:
            required_param_fields = [
                "gamma",
                "delta",
                "xm",
                "ym",
                "dh1",
                "dh2",
                "dm",
                "n",
                "mu",
                "bz",
                "lya",
                "wn0",
            ]

            missing = [
                field_name
                for field_name in required_param_fields
                if cleaned_data.get(field_name) in (None, "")
            ]
            if missing:
                raise ValidationError(
                    "Для параметрического тормоза должны быть заполнены все коэффициенты."
                )

            cleaned_data["parsed_force_curve_points"] = []
            cleaned_data["curve_file_provided"] = False
            cleaned_data["curve_source_brake_id"] = ""

        elif model_type == MagneticBrakeConfig.MODEL_TYPE_CURVE:
            uploaded_file = cleaned_data.get("force_curve_file")
            source_brake_id = (cleaned_data.get("curve_source_brake_id") or "").strip()
            catalog_source_id = cleaned_data.get("catalog_source_id")

            if uploaded_file:
                # Файл загружен прямо в форму
                parsed_points = self._parse_force_curve_file(uploaded_file)
                cleaned_data["parsed_force_curve_points"] = parsed_points
                cleaned_data["curve_file_provided"] = True
                cleaned_data["curve_source_brake_id"] = ""
            elif source_brake_id:
                # Reuse характеристики из ранее запущенного расчёта
                try:
                    int(source_brake_id)
                except ValueError as exc:
                    raise ValidationError(
                        "Некорректный идентификатор исходной характеристики тормоза."
                    ) from exc

                cleaned_data["parsed_force_curve_points"] = None
                cleaned_data["curve_file_provided"] = False
                cleaned_data["curve_source_brake_id"] = source_brake_id
            elif catalog_source_id:
                # Будет copy-on-use из каталога — фактическое чтение файла в view.
                # Здесь просто разрешаем сохранение без uploaded_file.
                cleaned_data["parsed_force_curve_points"] = None
                cleaned_data["curve_file_provided"] = False
                cleaned_data["curve_source_brake_id"] = ""
            else:
                raise ValidationError(
                    "Для тормоза, заданного графиком, необходимо загрузить Excel-файл "
                    "характеристики F(v) или выбрать тормоз из каталога."
                )

        return cleaned_data

    def _parse_force_curve_file(self, uploaded_file) -> list[dict]:
        self._validate_excel_extension(uploaded_file.name)

        try:
            uploaded_file.seek(0)
        except Exception:
            pass

        try:
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError(
                "Не удалось прочитать Excel-файл характеристики тормоза."
            ) from exc
        finally:
            try:
                uploaded_file.seek(0)
            except Exception:
                pass

        try:
            sheet = workbook.active
            parsed_points = self._parse_force_curve_sheet(sheet)
        finally:
            workbook.close()

        return parsed_points

    def _validate_excel_extension(self, filename: str) -> None:
        suffix = Path(filename).suffix.lower()
        allowed = {".xlsx", ".xlsm", ".xltx", ".xltm"}
        if suffix not in allowed:
            raise ValidationError(
                "Поддерживаются только Excel-файлы форматов .xlsx, .xlsm, .xltx, .xltm."
            )

    def _parse_force_curve_sheet(self, sheet) -> list[dict]:
        parsed_points: list[dict] = []
        header_skipped = False

        for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            velocity_raw = row[0] if len(row) > 0 else None
            force_raw = row[1] if len(row) > 1 else None

            if self._is_empty_cell(velocity_raw) and self._is_empty_cell(force_raw):
                continue

            if self._is_empty_cell(velocity_raw) or self._is_empty_cell(force_raw):
                raise ValidationError(
                    f"Строка {row_no}: должны быть заполнены и скорость, и сила."
                )

            velocity = self._coerce_excel_number(velocity_raw)
            force = self._coerce_excel_number(force_raw)

            # Допускаем одну строку заголовков в начале файла
            if velocity is None or force is None:
                if not header_skipped and not parsed_points:
                    header_skipped = True
                    continue
                raise ValidationError(
                    f"Строка {row_no}: значения скорости и силы должны быть числами."
                )

            if velocity < 0:
                raise ValidationError(
                    f"Строка {row_no}: скорость не может быть отрицательной."
                )

            if force < 0:
                raise ValidationError(
                    f"Строка {row_no}: сила торможения не может быть отрицательной."
                )

            parsed_points.append(
                {
                    "order": len(parsed_points) + 1,
                    "velocity": velocity,
                    "force": force,
                }
            )

        if len(parsed_points) < 2:
            raise ValidationError(
                "Для графического тормоза необходимо минимум 2 точки."
            )

        velocities = [point["velocity"] for point in parsed_points]

        if len(set(velocities)) != len(velocities):
            raise ValidationError(
                "Скорости в характеристике тормоза не должны повторяться."
            )

        if velocities != sorted(velocities):
            raise ValidationError(
                "Скорости в характеристике тормоза должны быть строго возрастающими."
            )

        return parsed_points

    @staticmethod
    def _is_empty_cell(value) -> bool:
        return value is None or (isinstance(value, str) and value.strip() == "")

    @staticmethod
    def _coerce_excel_number(value) -> float | None:
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            normalized = value.strip().replace(",", ".")
            if not normalized:
                return None
            try:
                return float(normalized)
            except ValueError:
                return None

        return None


class BaseMagneticBrakeFormSet(BaseFormSet):
    def clean(self):
        super().clean()

        if any(self.errors):
            return

        active_forms = 0
        for form in self.forms:
            if not hasattr(form, "cleaned_data"):
                continue
            if not form.cleaned_data:
                continue
            active_forms += 1

        if active_forms < 1:
            raise ValidationError("Необходимо задать минимум один тормоз.")


MagneticBrakeFormSet = formset_factory(
    MagneticBrakeForm,
    formset=BaseMagneticBrakeFormSet,
    extra=0,
    min_num=1,
    validate_min=True,
)


class CompareRunsForm(forms.Form):
    run_a = forms.ModelChoiceField(
        queryset=CalculationRun.objects.order_by("-created_at"),
        label="Расчёт A",
    )
    run_b = forms.ModelChoiceField(
        queryset=CalculationRun.objects.order_by("-created_at"),
        label="Расчёт B",
    )

    def clean(self):
        cleaned_data = super().clean()
        run_a = cleaned_data.get("run_a")
        run_b = cleaned_data.get("run_b")

        if run_a and run_b and run_a.pk == run_b.pk:
            raise ValidationError("Для сравнения нужно выбрать два разных расчёта.")

        return cleaned_data

# ============================================================================
# СРЕЗ 3a: Каталог тормозов
# ============================================================================

from .models import BrakeCatalog  # noqa: E402  (импорт здесь, чтобы не ломать сверху)


class BrakeCatalogForm(forms.ModelForm):
    """Форма создания/редактирования записи в каталоге тормозов."""

    class Meta:
        model = BrakeCatalog
        fields = [
            "name",
            "description",
            "model_type",
            "gamma", "delta", "n",
            "curve_file",
            # доп. параметры
            "xm", "ym", "dh1", "dh2", "dm", "mu", "bz", "lya", "wn0",
        ]
        widgets = {
            "description": forms.Textarea(attrs={"rows": 3}),
        }

    def clean(self):
        cleaned = super().clean()
        model_type = cleaned.get("model_type")

        if model_type == BrakeCatalog.MODEL_TYPE_PARAMETRIC:
            # Для параметрической модели все параметры обязательные.
            required_fields = [
                "gamma", "delta", "n",
                "xm", "ym", "dh1", "dh2", "dm",
                "mu", "bz", "lya", "wn0",
            ]
            missing: list[str] = []
            for field_name in required_fields:
                value = cleaned.get(field_name)
                if value is None or value == "":
                    missing.append(field_name)
                    # Отметим конкретное поле, чтобы рамка ошибки появилась рядом
                    self.add_error(
                        field_name,
                        "Обязательное поле для параметрической модели."
                    )
            if missing:
                # Общее сообщение наверх формы — какие поля не заполнены
                raise ValidationError(
                    "Для параметрической модели заполните все параметры. "
                    f"Не заполнены: {', '.join(missing)}."
                )

        elif model_type == BrakeCatalog.MODEL_TYPE_CURVE:
            # Для табличной модели нужен только файл F(v).
            curve_file = cleaned.get("curve_file")
            instance = getattr(self, "instance", None)
            already_has_file = bool(instance and instance.pk and instance.curve_file)
            if not curve_file and not already_has_file:
                self.add_error(
                    "curve_file",
                    "Обязательно загрузите файл F(v) для табличной модели."
                )
                raise ValidationError(
                    "Для табличной модели загрузите файл F(v)."
                )

        return cleaned

    def clean_name(self):
        """Имя должно быть уникальным (с учётом редактирования существующей записи)."""
        name = (self.cleaned_data.get("name") or "").strip()
        if not name:
            raise ValidationError("Имя обязательно.")

        qs = BrakeCatalog.objects.filter(name=name)
        if self.instance and self.instance.pk:
            qs = qs.exclude(pk=self.instance.pk)
        if qs.exists():
            raise ValidationError("Запись с таким именем уже существует в каталоге.")
        return name
